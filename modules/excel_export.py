import io
import pandas as pd
from openpyxl import Workbook
from .formatting import style_sheet
from .summaries import customer_summary, batch_summary

def write_df(ws, df):
    for ci, col in enumerate(df.columns, start=1):
        ws.cell(1, ci, col)
    for ri, row in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row, start=1):
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None
            ws.cell(ri, ci, val)

def build_workbook(df_profit, df_czasy=None, df_kliki=None, df_farby=None,
                   rates=None, click_costs=None, prepress=None,
                   params=None, machine_cols=None, selected_month="Wszystkie"):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Profitability")
    write_df(ws, df_profit)
    style_sheet(ws, df_profit, machine_cols)

    if df_czasy is not None:
        ws2 = wb.create_sheet("czasy")
        out = df_czasy.drop(columns=[c for c in ["_zp", "_machine", "_minutes", "_rate"] if c in df_czasy.columns], errors="ignore")
        write_df(ws2, out)
        style_sheet(ws2, out)

    if df_kliki is not None:
        ws3 = wb.create_sheet("Kliki")
        write_df(ws3, df_kliki)
        style_sheet(ws3, df_kliki)

    if df_farby is not None:
        ws4 = wb.create_sheet("Farby Offset")
        write_df(ws4, df_farby)
        style_sheet(ws4, df_farby)

    df_rates = pd.DataFrame(list((rates or {}).items()), columns=["Nazwa maszyny", "Stawka rbg"])
    ws5 = wb.create_sheet("Stawki")
    write_df(ws5, df_rates)
    style_sheet(ws5, df_rates)

    cc_rows = []
    for press, colors in (click_costs or {}).items():
        for color, cost in colors.items():
            cc_rows.append({"Press Name": press, "Color": color, "Koszt": cost})
    df_cc = pd.DataFrame(cc_rows)
    ws6 = wb.create_sheet("Koszty klików")
    write_df(ws6, df_cc)
    style_sheet(ws6, df_cc)

    pp_rows = [{"Klient": k, "Stawka Digital": v.get("digital", 10), "Stawka Offset": v.get("offset", 40)} for k, v in (prepress or {}).items()]
    df_pp = pd.DataFrame(pp_rows)
    ws7 = wb.create_sheet("Prepress")
    write_df(ws7, df_pp)
    style_sheet(ws7, df_pp)

    df_params = pd.DataFrame([params or {}]).T.reset_index()
    df_params.columns = ["Parametr", "Wartość"]
    ws8 = wb.create_sheet("Parametry")
    write_df(ws8, df_params)
    style_sheet(ws8, df_params)

    df_sum = customer_summary(df_profit, selected_month)
    ws9 = wb.create_sheet("Podsumowanie")
    write_df(ws9, df_sum)
    style_sheet(ws9, df_sum)

    df_batch = batch_summary(df_profit, selected_month)
    ws10 = wb.create_sheet("Batch summary")
    write_df(ws10, df_batch)
    style_sheet(ws10, df_batch)

    ws11 = wb.create_sheet("Kokpit")
    write_df(ws11, df_sum)
    style_sheet(ws11, df_sum)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
