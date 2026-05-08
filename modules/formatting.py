from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILL_HEADER = PatternFill("solid", fgColor="3A0000")
FILL_DL = PatternFill("solid", fgColor="C5DCF5")
FILL_MAT = PatternFill("solid", fgColor="C5F5D9")
FONT_WHITE = Font(bold=True, color="FFFFFF")
FONT_NORMAL = Font(color="1F1F1F")
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HIDDEN_COLS = {
    "kontrbigi [29]", "duplex [30]", "offset [31]", "add on evey box [32]",
    "autobottom / ekstra lim [33]", "pantone [34]", "e-flute [35]",
    "item number inlay [36]", "pod price [37]", "clames [38]", "transport [39]",
    "kooperacja [40]", "b2 price [41]", "b1 price [42]", "energia [43]",
    "fix price [44]", "cena tkw [45]", "click price [46]",
    "płyty lakierujące [21]", "matryca braille- grawer [22]",
    "patryca braille- wewnatrz [23]"
}

MATERIAL_COLS = {"Papier [16]", "Klej [17]", "Lakiery [20]", "Opakowania zbiorcze [24]",
                 "Other Materials", "Offset inks", "Płyta offsetowa", "Kliki final", "Total Materials"}

def style_sheet(ws, df, machine_cols=None):
    machine_cols = set(machine_cols or [])
    dl_cols = machine_cols | {"Prepress costs", "Total DL"}
    for ci, col in enumerate(df.columns, start=1):
        cell = ws.cell(1, ci)
        cell.font = FONT_WHITE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        if col in dl_cols:
            cell.fill = FILL_DL
            cell.font = Font(bold=True, color="000000")
        elif col in MATERIAL_COLS:
            cell.fill = FILL_MAT
            cell.font = Font(bold=True, color="000000")
        else:
            cell.fill = FILL_HEADER

        if col.lower().strip() in HIDDEN_COLS:
            ws.column_dimensions[get_column_letter(ci)].hidden = True

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.font = FONT_NORMAL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for ci, col in enumerate(df.columns, start=1):
        width = max(10, min(40, len(str(col)) + 2))
        for r in range(2, min(ws.max_row, 50) + 1):
            val = ws.cell(r, ci).value
            if val is not None:
                width = max(width, min(40, len(str(val)) + 2))
        ws.column_dimensions[get_column_letter(ci)].width = width
